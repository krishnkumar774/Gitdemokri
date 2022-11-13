import openpyxl
Dist={}
book=openpyxl.load_workbook("C:\\prc\\pyhonDemoData.xlsx")
sheet =book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range (2,sheet.max_column+1):
            Dist[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(Dist)